import pickle

from openpyxl import load_workbook

from scrapper.model.schedule import Area
from scrapper.walker import logger
from scrapper.walker.metadator import MetadataScanner

import log_config


def get_input_file_names():
    import glob
    names = glob.glob('./resources/excel/*.xlsx')
    return names


def right_number_of_days(month_no, no_of_days):
    MONTH_DAYS = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    against = MONTH_DAYS[month_no - 1]

    if type(against) == list:
        return no_of_days in against
    else:
        return no_of_days == against


def parse_all():
    names = get_input_file_names()
    # names = ['./resources/excel/bojano-2-harmonogram-2021-1-1.xlsx']
    # names = ['./resources/excel/kamien-kowalewo-ul.-zimowa-harmonogram-2021-.xlsx']
    # names = ['./resources/excel/czestkowo-glazica-szemud-ulica-bursztynowa-donimierz-ulica-zablotna-maszyna-harmonogram-2021-kalendarz-rejon-2.xlsx']
    # names = ['./resources/excel/bedargowo-zeblewo-harmonogram-2021-1-1.xlsx']
    # names = ['./resources/excel/bojano-3-harmonogram-2021.xlsx']

    result = []

    for name in names:
        path = name
        logger.info(f"path: ----------------------------------------- {path}")
        print(f"path: ----------------------------------------- {path}")

        wb = load_workbook(path)
        ws = wb.worksheets[0]
        # print(f"Sheet {ws} max rows: {ws.max_row} max cols: {ws.max_column}")

        metadator = MetadataScanner(ws)
        area_name, streets = metadator.extract_area(), metadator.extract_streets()

        zurlowany_path = path.replace("./resources/excel/", "http://szemud.pl/wp-content/uploads/2015/10/")
        zurlowany_path = zurlowany_path.replace(".xlsx", ".pdf")

        result.append(Area(area_name, streets.replace('\n', " "), zurlowany_path))

    return result


result = parse_all()
pickle.dump(result, open("areas.p", "wb"))

# try:
#     areas = pickle.load(open("areas.p", "rb"))
# except Exception:
#     logger.error("Error reading the schedule from pickled file `areas.p` (produced by `test_extract_metadata.py`")
#
# print("Areas:")
# print(areas)
# print("Events  :")
# print(schedule)
#
# for a in areas:
#     if not a.name or not a.description:
#         print("name")
#         print(a.name)
#         print("desc")
#         print(a.description)
#         print("source")
#         print(a.source)
