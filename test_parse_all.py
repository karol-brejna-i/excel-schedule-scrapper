import pickle

from openpyxl import load_workbook

from scrapper.env_config import CURRENT_SCHEDULE
from scrapper.model.schedule import Area
from scrapper.walker import logger
from scrapper.walker.metadator import MetadataScanner
from scrapper.walker.scanner import TableScanner

import log_config


def get_input_file_names():
    import glob
    names = glob.glob('./resources/excel/*.xlsx')
    return names


def parse_new():
    year = CURRENT_SCHEDULE
    names = get_input_file_names()
    # names = ['./resources/excel/bojano-2-harmonogram-2021-1-1.xlsx']
    # names = ['./resources/excel/kamien-kowalewo-ul.-zimowa-harmonogram-2021-.xlsx']
    # names = ['./resources/excel/czestkowo-glazica-szemud-ulica-bursztynowa-donimierz-ulica-zablotna-maszyna-harmonogram-2021-kalendarz-rejon-2.xlsx']
    # names = ['./resources/excel/bedargowo-zeblewo-harmonogram-2021-1-1.xlsx']
    # names = ['./resources/excel/bojano-3-harmonogram-2021.xlsx']

    areas, schedules = [], []
    for name in names:
        path = name
        logger.info(f"path: ----------------------------------------- {path}")

        wb = load_workbook(path)
        ws = wb.worksheets[0]
        # print(f"Sheet {ws} max rows: {ws.max_row} max cols: {ws.max_column}")
        scanner = TableScanner(ws)

        # extract metadata (area name, addresses)
        metadator = MetadataScanner(ws)
        area_name, streets = metadator.extract_area(), metadator.extract_streets()

        zurlowany_path = path.replace("./resources/excel/", "http://szemud.pl/files/file/Formularze_Do_pobrania/Gospodarka_Odpadami/Odpady2012-harmonogram/")
        zurlowany_path = zurlowany_path.replace(".xlsx", ".pdf")

        area = Area(area_name, streets.replace('\n', " "), zurlowany_path)
        areas.append(area)
        # extract event data
        schedule = scanner.get_schedule(year, area_name, zurlowany_path)

        schedules.append(schedule)

    return areas, schedules


result = parse_new()

#
# print("results------------")
# print(result[0])
# print(result[1])
# schedules = result[1]
# s = schedules[0]
# print(s.months)
# for m in s.months.items():
#     print(m)
#
pickle.dump(result[0], open("areas.p", "wb"))
pickle.dump(result[1], open("schedule.p", "wb"))
