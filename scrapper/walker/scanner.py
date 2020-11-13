from openpyxl.cell import MergedCell
from openpyxl.styles.borders import BORDER_NONE

from scrapper.model.schedule import Calendar
from scrapper.calendar.parsers import CalendarDataParser
from scrapper.model.cells import MonthCellsDescription, DayData, DayCellDescription
from scrapper.walker import logger
from scrapper.walker.walker import Walker, MONTH_NAMES

DAY_NAMES = ['PN', 'WT', 'ŚR', 'CZ', 'PT', 'SO', 'ND']

BORDER_DIRECTIONS = ['left', 'right', 'top', 'bottom']

COLOR_NAMES_LOOKUP = {'00000000': '',
                      'FFFF0000': 'RED',
                      'FF6FAC46': 'GREEN', 'FF6FAD46': 'GREEN',
                      'FF00AFEF': 'BLUE', 'FF5B9BD4': 'BLUE',
                      'FFFFFF00': 'YELLOW',
                      'FF000000': 'BLACK', 'FF0D0D0D': 'BLACK',
                      'FFC55A11': 'BROWN', 'FF00B0F0': 'BROWN', 'FFC65810': 'BROWN',
                      'FF833B0B': "BLACK AND BROWN", 'FF843B0C': "BLACK AND BROWN"}


class TableScanner:
    """
    Walks through worksheet cells and tries to determine size of sections responsible for holding data for given day
    """

    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.walker = Walker(worksheet)

    def check_if_merged(self, row, col, limit_col, limit_row=999):
        if col >= limit_col or row >= limit_row:
            logger.debug("out of col limits")
            result = False
        else:
            neighbour_cell = self.worksheet.cell(row, col)
            logger.debug(f"neighbour {neighbour_cell}, type: {type(neighbour_cell)}")

            result = type(neighbour_cell) == MergedCell
        return result

    # TODO give more meaningful name
    @classmethod
    def check_if_none(cls, border_value):
        """
        checks if given border_value (an atribute of a cell) is None (no border)
        :param border_value:
        :return:
        """
        return border_value.style == BORDER_NONE

    def get_borders(self, cell):
        """
        Get left, top, right, bottom values of borders for given cell.
        :param cell:
        :return:
        """
        border = cell.border

        result = []
        for attr in BORDER_DIRECTIONS:
            if not self.check_if_none(getattr(border, attr)):
                result.append(attr)
        return result

    def walk_through_row(self, starting_row, starting_col, limit_col=20, limit_row=999):
        logger.debug(f"walk_through_row {starting_row}, {starting_col}, {limit_col}, {limit_row}")

        col = starting_col
        row = starting_row

        # XXX TODO big assumption: we start in upper left cell of the calendar
        # right after the row with day names

        result = []

        finish = False
        no_of_days = 0
        while not finish:
            cell = self.worksheet.cell(row=row, column=col)

            logger.debug(f"checking {col}:   {self.walker.get_cell_info(cell)}")

            # get the borders of given cell (left, top, bottom, right)
            borders = self.get_borders(cell)
            logger.debug(f"cell borders: {borders}")

            # for now, we assume that 2x2 is the biggest possible section, so no real traversal right and down is done...
            width = 1

            seek_further = True
            while seek_further:

                border_condition = 'right' not in borders
                merged_condition = self.check_if_merged(row, col + width, limit_col, limit_row)

                if border_condition or merged_condition:
                    # if 'right' not in borders or self.check_if_merged(row, col + 1, limit_col, limit_row):
                    width = width + 1
                else:
                    seek_further = False

            height = 1
            border_condition = 'bottom' not in borders
            merged_condition = self.check_if_merged(row + 1, col, limit_col, limit_row)
            if border_condition or merged_condition:
                # if 'bottom' not in borders or self.check_if_merged(row + 1, col, limit_col, limit_row):
                height = height + 1

            value = cell.value

            # skip to the next cell
            col = col + width
            # remember the data for this day
            result.append((value, (height, width), cell.coordinate))
            no_of_days = no_of_days + 1

            if col > limit_col or no_of_days >= 7:
                finish = True

        return result

    def walk_through_row_with_column_widths(self, starting_row, starting_col, column_widths):
        logger.debug(f"walk_through_row {starting_row}, {starting_col}")

        limit_col, limit_row = 999, 999

        col = starting_col
        row = starting_row

        # XXX TODO big assumption: we start in upper left cell of the calendar
        # right after the row with day names

        result = []

        finish = False
        no_of_days = 0
        while not finish:
            cell = self.worksheet.cell(row=row, column=col)

            logger.debug(f"checking {col}:   {self.walker.get_cell_info(cell)}")

            # get the borders of given cell (left, top, bottom, right)
            borders = self.get_borders(cell)
            logger.debug(f"cell borders: {borders}")

            # for now, we assume that 2x2 is the biggest possible section, so no real traversal right and down is done...
            width = column_widths[no_of_days]
            height = 1

            border_condition = 'bottom' not in borders
            merged_condition = self.check_if_merged(row + 1, col, limit_col, limit_row)
            if border_condition or merged_condition:
                # if 'bottom' not in borders or self.check_if_merged(row + 1, col, limit_col, limit_row):
                height = height + 1

            value = cell.value

            # skip to the next cell
            col = col + width
            # remember the data for this day
            result.append((value, (height, width), cell.coordinate))
            no_of_days = no_of_days + 1

            if col > limit_col or no_of_days >= 7:
                finish = True

        return result

    # TODO move to Walker
    def find_months_in_a_column(self, starting_row, starting_col, limit_row=100):
        logger.debug(f"--- find_months_in_a_column({starting_row},{starting_col}, limit_row=100):")

        column_months = []
        # find month's calendar sections for given column
        month_info = self.walker.seek_down(starting_row, starting_col, MONTH_NAMES, self.worksheet.max_row)
        column_months.append(month_info)

        month_info = self.walker.seek_down(month_info[0] + 6, month_info[1], MONTH_NAMES, self.worksheet.max_row)
        column_months.append(month_info)

        # determine length of the sections
        # month_sections = []
        month_descriptions = []
        for month_info in column_months:
            sizes = self.walker.find_right_ending(month_info[0], month_info[1])
            # determine column widths
            day_cells = self.walk_through_row(month_info[0] + 1, month_info[1], sizes[1])
            widths = list(map(lambda d: d[1][1], day_cells))
            month_description = MonthCellsDescription(month_info[2], (month_info[0], month_info[1]),
                                                      (sizes[0], sizes[1]), widths)
            month_descriptions.append(month_description)
        return month_descriptions

    # TODO probably also a walker responsibility
    def find_next_column(self, next_coords):
        ending_row, ending_col = next_coords[0], next_coords[1]
        # find next column cell
        next_month = self.walker.seek_right(ending_row, ending_col + 1, MONTH_NAMES)
        next_column = self.find_months_in_a_column(next_month[0], next_month[1])
        return next_column

    # TODO Walker?
    def find_months(self, starting_row=10, starting_col=1):
        result = []
        # determine left-most months in the calendar
        months = self.find_months_in_a_column(starting_row, starting_col)
        result.extend(months)

        # second column
        next_coords = months[0].top_right_cell
        months = self.find_next_column(next_coords)
        result.extend(months)

        # third column
        next_coords = months[0].top_right_cell
        months = self.find_next_column(next_coords)
        result.extend(months)

        return result

    def scan_month(self, month):
        starting_row, starting_col, ending_col = month.top_left_cell[0], month.top_left_cell[1], month.top_right_cell[1]
        logger.debug(f"scan month  {month}      starting row, col = {starting_row}, {starting_col}")

        # XXX TODO hardcoded. is there a better way for doing this
        # the first row is a row with a month name
        # the second row is a row with day names
        # data starts from third row
        starting_row += 2
        result = []
        for i in range(1, 7):
            # get data for row
            # row_data = self.walk_through_row(starting_row, starting_col, ending_col)
            row_data = self.walk_through_row_with_column_widths(starting_row, starting_col, month.column_widths)

            # remember the result
            result.append(row_data)

            # advance to the next row (cell row number depends on the highest cell in previous row)
            step = max(row_data, key=lambda cl: cl[1][0])[1][0]
            #
            # step = map(lambda x: x[1][0], max(row_data, key=lambda cl: cl[1][0]))
            starting_row = starting_row + step

        return result

    def get_month_day_cells(self, month):
        month_data = self.scan_month(month)
        flattened = [item for row in month_data for item in row]
        filtered = list(filter(lambda x: x[0], flattened))
        result = list(map(lambda x: DayCellDescription(x[0], coords=x[2], size=x[1]), filtered))
        return result

    def get_data_for_day_cell(self, day_descriptor: DayCellDescription):
        logger.debug(f"get_data_for_day(self, {day_descriptor}):")

        cell = self.worksheet[day_descriptor.coords]
        starting_row = cell.row
        starting_col = cell.column

        neighbour_cells = set()
        for x in range(0, day_descriptor.size[1]):
            for y in range(0, day_descriptor.size[0]):
                cell = self.worksheet.cell(starting_row + y, starting_col + x)
                neighbour_cells.add(cell)

        # XXX TODO debug
        for p in neighbour_cells:
            if p.fill.fgColor.rgb not in COLOR_NAMES_LOOKUP:
                logger.warn(f"dziwny kolor: {p.fill.fgColor.rgb}          {p} ")

        colors = set(
            map(lambda p: COLOR_NAMES_LOOKUP[
                p.fill.fgColor.rgb] if p.fill.fgColor.rgb in COLOR_NAMES_LOOKUP else 'UNKNOWN',
                neighbour_cells
                )
        )
        colors.discard('')
        return DayData(day_descriptor.value, colors)

    def get_data_for_day_cells(self, day_descriptors: list):
        result = list(map(lambda x: self.get_data_for_day_cell(x), day_descriptors))
        return result

    # XXX TODO introduce leap year logic
    @staticmethod
    def get_number_of_days(year, month_no):
        MONTH_DAYS = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        return MONTH_DAYS[month_no - 1]

    @staticmethod
    def right_number_of_days(year, month_no, no_of_days):
        against = TableScanner.get_number_of_days(year, month_no)

        if type(against) == list:
            return no_of_days in against
        else:
            return no_of_days == against

    def get_calendar(self, year, area_name, path):
        # find cells for given months
        months = self.find_months()

        # create calendar object
        calendar = Calendar(year, area_name, path)

        for month in months:
            month_name = month.month_name
            month_no = MONTH_NAMES.index(month_name) + 1
            logger.info(f"### {month_name}, {month_no}")
            day_cells = self.get_month_day_cells(month)
            days_data = self.get_data_for_day_cells(day_cells)
            no_of_days = len(day_cells)

            try:
                # XXX TODO introduce leap year logic
                m = CalendarDataParser.month_from_day_data(month_no, self.get_number_of_days(year, month_no), days_data)
                calendar.set_month(m.month_of_year, m)
            except Exception as e:
                logger.error(f"------------ {month} z {path} jest nieudany. przyjrzyj się.")
                logger.error(e)

            if not self.right_number_of_days(year, month_no, no_of_days):
                logger.warning(f"!!!!!!!!!!!!!!!!!!!!!!!!! dziwny dziwny; {path} {month_name} {month_no} {no_of_days}")

        return calendar
