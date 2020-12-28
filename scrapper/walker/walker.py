from openpyxl.cell import Cell

from scrapper.env_config import MONTH_NAMES
from scrapper.walker import logger


class Walker:
    def __init__(self, worksheet):
        self.worksheet = worksheet

    @classmethod
    def get_cell_info(cls, cell):
        if type(cell) == Cell:
            return f"cell({cell.column_letter}{cell.row}) --- {cell.value}--- {cell.has_style}"
        else:
            return f"mergedCell({cell.column},{cell.row}) --- {cell.value}--- {cell.has_style}"

    # TODO arguments are row, col -- but when presenting cell info we show col, row; I think the latter is the convention
    def seek_down(self, starting_row, starting_col, value_to_find, limit_row=100):
        """
        Starts traversing cells down (row by row) from given position (starting_row, starting_col)
        trying to find specific value (value_to_find).
        :param starting_row:
        :param starting_col:
        :param value_to_find:
        :param limit_row: stop searching after looking at this many rows (the limit)
        :return: (row, col, value); if not found value is none
        """
        logger.debug(f'seek_down({starting_row}, {starting_col}, {value_to_find}, {limit_row}')
        col = starting_col
        row = starting_row
        value = None
        for row in range(starting_row, starting_row + limit_row):
            cell = self.worksheet.cell(row=row, column=col)
            if type(value_to_find) == list:
                if cell.value in value_to_find:
                    value = cell.value
                    break
            else:
                if cell.value == value_to_find:
                    value = cell.value
                    break
        # TODO maybe should return None if the value is not found?
        return row, col, value

    def seek_right(self, starting_row, starting_col, value_to_find, limit_col=100):
        """
        Starts traversing cells right (column by column) from given position (starting_row, starting_col)
        trying to find specific value (value_to_find).

        :param starting_row:
        :param starting_col:
        :param value_to_find:
        :param limit_col: stop searching after looking at this many columns (the limit)
        :return:
        """
        col = starting_col
        row = starting_row
        value = None
        for col in range(starting_col, starting_col + limit_col):
            cell = self.worksheet.cell(row=row, column=col)
            logger.debug(self.get_cell_info(cell))
            if type(value_to_find) == list:
                logger.debug("checking value against a list")
                if cell.value in value_to_find:
                    value = cell.value
                    break
            else:
                logger.debug("checking specific")
                if cell.value == value_to_find:
                    value = cell.value
                    break

        # TODO maybe should return None if the value is not found?
        return row, col, value

    def find_right_ending(self, starting_row, starting_col, limit_col=30, limit_row=30):
        logger.debug(f"find_right_ending({starting_row}, {starting_col}, {limit_col}, {limit_row})")

        # step 1: is does the entry cell contain a month name?
        starting_cell = self.worksheet.cell(starting_row, starting_col)
        if starting_cell.value in MONTH_NAMES:
            row = starting_row
            # the cell with month's name is merged, find the end of merged cell
            for col in range(starting_col + 1, starting_col + limit_col):
                cell = self.worksheet.cell(row=row, column=col)
                if type(cell) == Cell:
                    result = (row, col - 1)
                    break
        else:
            logger.debug("wrong starting point!")
            result = None

        return result

    @classmethod
    def value_is_a_months_day(cls, value):
        numbers = [i for i in range(1, 32)]
        return value in numbers
